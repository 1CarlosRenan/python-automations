import openpyxl

# Criar uma planilha (book)
book = openpyxl.Workbook()
# Como visualizar páginas existentes
print(book.sheetnames)
# Como criar uma página
book.create_sheet('Frutas')
# Como selecionar uma página
frutas_page = book['Frutas']
# Colocar valores em linha
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Fruta 2', '2', 'R$5,90'])
frutas_page.append(['Fruta 3', '10', 'R$2,90'])
frutas_page.append(['Fruta 4', '2', 'R$31,90'])

# Salvar a planilha
book.save('Planilha de Compras.xlsx') 
