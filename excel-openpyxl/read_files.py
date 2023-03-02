import openpyxl

# Carregando arquivo
book = openpyxl.load_workbook('Planilha de Compras.xlsx')

# Selecionando Página
frutas_page = book['Frutas']

# Imprimindo os dados de cada linha
for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        print(cell.value)

# Alterando valor com condicional
for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        if cell.value == 'Banana':
            cell,value = 'Fruta 1'

# Salvar as alterações
book.save('Planilha de Compras v2.xlsx')